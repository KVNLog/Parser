import json
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from  .ExcelSaver import ExcelSaver


class ExcelSaverZakupki(ExcelSaver):

    def _create_json_sheet(self):
        """Создаёт отдельный лист для каждого артикула и записывает соответствующие данные."""
        from src.logger import parser_logger
        try:
            # Получаем последний JSON-файл
            self.json_file = self._get_latest_json(self.json_folder)
            parser_logger.info(f"{self.__class__.__name__}: Создание листов из JSON-файла '{self.json_file}'")

            # Загружаем данные из JSON
            self.data = self._load_price_from_json()

            # Проверяем, есть ли данные в JSON
            if not self.data or "Данные" not in self.data:
                parser_logger.warning(f"{self.__class__.__name__}: В JSON-файле '{self.json_file}' отсутствуют данные.")
                return

            # Создаём новый Workbook
            self.workbook = Workbook()
            first_sheet = self.workbook.active
            first_sheet.title = "Артикулы"

            # Записываем все артикулы в первый столбец первого листа
            for row_index, article in enumerate(self.articles, start=1):
                first_sheet.cell(row=row_index, column=1, value=article)

            # Проходим по каждому артикулу
            for article in self.articles:
                # Определяем имя листа
                sheet_name = str(article)[:31]  # Ограничение на длину имени листа в Excel
                parser_logger.debug(f"{self.__class__.__name__}: Создание листа для артикула: {sheet_name}")

                # Проверяем, существует ли лист с таким именем, и удаляем его, если да
                if sheet_name in self.workbook.sheetnames:
                    parser_logger.warning(f"{self.__class__.__name__}: Лист '{sheet_name}' уже существует, удаляем его")
                    self.workbook.remove(self.workbook[sheet_name])

                # Создаём новый лист
                ws = self.workbook.create_sheet(title=sheet_name)

                # Записываем заголовки
                headers = ["description", "url", "price"]
                ws.append(headers)

                # Ищем данные для текущего артикула
                article_found = False
                for item in self.data.get("Данные", []):
                    if article in item:
                        article_data = item[article]
                        row = [
                            article_data.get("description", "Нет описания"),
                            article_data.get("url", "Нет URL"),
                            article_data.get("price", "Нет цены")
                        ]
                        ws.append(row)
                        article_found = True

                if not article_found:
                    parser_logger.debug(f"{self.__class__.__name__}: Для артикула '{article}' данные не найдены")
                    ws.append(["Данные не найдены", "", ""])

        except Exception as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка при создании листов для артикулов: {e}")

    def process_data(self):
        """Выполняет полный цикл обработки данных: загрузка, обновление и сохранение."""
        from src.logger import parser_logger
        try:
            parser_logger.info(f"{self.__class__.__name__}: Начало обработки данных")

            self._open_excel()
            parser_logger.info(f"{self.__class__.__name__}: Excel-файл успешно загружен")

            self._create_json_sheet()
            parser_logger.info(f"{self.__class__.__name__}: Новый лист с JSON-данными успешно создан")

            self._save_to_excel()
            parser_logger.info(f"{self.__class__.__name__}: Изменения сохранены в Excel-файл")

            parser_logger.info(f"{self.__class__.__name__}: Обработка данных завершена")

        except Exception as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка во время обработки данных: {e}")


# # Пример использования
# saver = ExcelSaver(json_folder="JSON/ChipDipData")
# saver1 = ExcelSaver(json_folder="JSON/eBayData")
# saver2 = ExcelSaver(json_folder="JSON/ETMData")
# saver3 = ExcelSaver(json_folder="JSON/YandexMarketData")
# saver.process_data()
# saver1.process_data()
# saver2.process_data()
# saver3.process_data()
# saver3.aggregate_prices_to_first_sheet()
