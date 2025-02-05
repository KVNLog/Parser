import json
import os
import pandas as pd
from openpyxl import load_workbook, Workbook


class ExcelSaver:
    def __init__(self, excel_file="Рабочий.xlsx", json_folder="json_data"):
        self.excel_file = excel_file  # Фиксированный путь к Excel-файлу
        self.json_folder = json_folder

        self.df = None  # DataFrame для работы с Excel
        self.workbook = None  # Workbook для работы с несколькими листами
        self.articles = []  # Список артикулов из первого столбца

    def _get_latest_json(self, folder):
        """Находит самый свежий JSON-файл в папке."""
        json_files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith(".json")]
        if not json_files:
            raise FileNotFoundError("Нет JSON-файлов в указанной папке.")
        latest_file = max(json_files, key=os.path.getmtime)
        print(f"Выбран самый свежий JSON-файл: {latest_file}")
        return latest_file

    def _load_price_from_json(self):
        """Загружает данные из JSON-файла."""
        self.json_file = self._get_latest_json(self.json_folder)  # Выбираем самый свежий JSON-файл
        try:
            with open(self.json_file, 'r', encoding='utf-8') as file:
                print("JSON данные загружены.")
                return json.load(file)
        except FileNotFoundError:
            print("Файл JSON не найден, данные не загружены.")
            return {}

    def _open_excel(self):
        """Открывает существующий Excel-файл и загружает артикулы с первого листа."""
        try:
            self.workbook = load_workbook(self.excel_file)
            sheet_name = self.workbook.sheetnames[0]  # Первый лист
            self.df = pd.read_excel(self.excel_file, sheet_name=sheet_name, engine='openpyxl', header=None)
            self.articles = self.df.iloc[:, 0].dropna().astype(str).tolist()  # Загружаем артикулы
            print("Excel файл загружен, артикулы считаны.")
        except FileNotFoundError:
            print("Файл не найден, процесс остановлен.")
            raise

    def _create_json_sheet(self):
        """Создаёт новый лист с именем JSON-файла и записывает данные по артикулам."""
        self.json_file = self._get_latest_json(self.json_folder)  # Выбираем самый свежий JSON-файл
        self.data = self._load_price_from_json()  # Загружаем JSON при создании объекта
        sheet_name = os.path.basename(self.json_file).split('.')[0].split("_")[0]  # Имя нового листа
        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])
        self.workbook.create_sheet(sheet_name)
        ws = self.workbook[sheet_name]
        # ws.append(["Артикул", "Цена 1", "Цена 2", "Цена 3", "..."])  # Заголовки для нескольких цен

        for article in self.articles:
            prices = []
            for item in self.data.get("Данные", []):
                if article in item:
                    price = item[article].get("price")
                    prices.append(price)
            if prices:
                ws.append([article] + prices)  # Каждая цена в отдельной ячейке
            else:
                ws.append([article, "Не найдено"])

        print(f"Создан новый лист '{sheet_name}' с данными из JSON.")

    def _save_to_excel(self):
        """Сохраняет изменения в Excel-файл."""
        self.workbook.save(self.excel_file)
        print("Изменения сохранены в Excel.")

    def process_data(self):
        """Выполняет полный цикл обработки данных: загрузка, обновление и сохранение."""
        self._open_excel()
        self._create_json_sheet()
        self._save_to_excel()
        print("Обработка данных завершена.")

    def aggregate_prices_to_first_sheet(self):
        """Собирает все цены с каждого листа и аккумулирует их на первом листе."""
        self._open_excel()
        first_sheet_name = self.workbook.sheetnames[0]
        first_sheet = self.workbook[first_sheet_name]

        # Словарь для аккумулирования цен по артикулам
        aggregated_data = {article: [] for article in self.articles}

        # Обход всех листов, кроме первого
        for sheet_name in self.workbook.sheetnames[1:]:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, values_only=True):
                article = row[0]
                prices = row[1:]
                if article in aggregated_data:
                    aggregated_data[article].extend([price for price in prices if price])

        # # Очистка первой страницы от старых данных
        first_sheet.delete_rows(1, first_sheet.max_row)

        # Запись данных на первый лист
        # first_sheet.append(["Артикул", "Цена 1", "Цена 2", "Цена 3", "..."])
        for article, prices in aggregated_data.items():
            first_sheet.append([article] + prices)

        # Сохранение изменений
        self._save_to_excel()

        print("Цены аккумулированы на первом листе и файл сохранён.")


# # Пример использования
# saver = ExcelSaver(json_folder="JSON/ChipDipData")
# saver1 = ExcelSaver(json_folder="JSON/eBayData")
# saver2 = ExcelSaver(json_folder="JSON/ETMData")
# saver3 = ExcelSaver(json_folder="JSON/YandexMarketData")
# saver.process_data()
# saver1.process_data()
# saver2.process_data()
# saver3.process_data()
# saver3.aggregate_prices_to_first_sheet()
